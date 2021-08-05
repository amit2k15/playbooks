import pymysql as MySQLdb
import csv
import time
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime

#db = MySQLdb.connect("192.168.255.187", "reportuser", "reportuser","sapdb_new")
db = MySQLdb.connect("127.0.0.1", "amit", "amit@123","sapdb_new")

from_time=1579717800
to_time=1582568940
dir_name=datetime.fromtimestamp(from_time).strftime("%B %Y")
if(not os.path.exists(dir_name)):
	os.mkdir(dir_name)
cursor = db.cursor()
#groupids_list=[107,1021,48,59,78,100,86,947,69,541,298,81,123,127,58,44,73,74,11,104,105,70,47,63,239,61,71,181,67,72,94,97,31,132,41,88,89,64,147,148,82,154,149,152,153,150,141,145,146,161,163,162,159,158,156,160,151,170,186,192,176,302,165,173,312,198,174,175,188,294,348,353,356,369,379,382,371,349,384,376,418,399,383,377,337,345,299,467,344,409,465,415,408,350,410,470,468,475,411,599,413,558,609,223,611,374,614,559,612,568,372,402,656,639,654,676,660,610,652,589,684,677,664,675,381,665,685,689,659,700,725,752,751,236,531,401,759,772,755,794,692,778,694,653,476,222,712,797,789,803,856,799,862,858,881,897,905,902,868,885,865]
#groupids_list=[1021,1036,222,970,856,957,284,794,204,748,285,141,972,1016,1017,1018,1034,224,535,737,300,350,742,744,550,719,726,858,862,946,902,222,735,905,244,301,539,730,71,239,263,266,308,536,268,517,732,523,553,252,533,611,246,528,287,209,215,262,218,279,278,277,682,778,882,869,260,303,212,210,543,529,524,656,727,219,531,249,267,545,538,749,522,198,838,304,261,236,272,202,230,254,718,721,803,789,530,613,561,208,220,271,286,214,207,213,298,265,299,541,527,519,560,516,747,612,773,734,722,723,807,745,653,754,797,799,885,654,761,720,418,614,692,223,251,257,221,275,518,534,958,719,639,724,741,609,930,873,914,985,888,909,897,881,947,1002,865,949,940,974,977,919]
groupids_list=[773]
to_avoid_disks=['C:','/','/boot','/home','/usr','/tmp','/var']
consider_disks=['sap','hana','sybase','softs']
for groupids in groupids_list:
	hostsinfo=[]
	query="select * from groups where groupid="+str(groupids)
	cursor.execute(query)
	groups = cursor.fetchall()
	groupname=""
	for group in groups:
		groupname=group[1]
	query="select h.hostid,h.host from hosts_groups hg join hosts h on hg.hostid=h.hostid where h.status=0 and hg.groupid="+str(groupids)
	cursor.execute(query)
	hosts = cursor.fetchall()
	query="select h.hostid,h.ip from hosts_groups hg join interface h on hg.hostid=h.hostid where hg.groupid="+str(groupids)
	cursor.execute(query)
	ips = cursor.fetchall()
	# query="select h.hostid,round(avg(t.VALUE_AVG),3)  from trends t join items i on t.ITEMID=i.ITEMID join functions f on f.ITEMID=i.ITEMID join triggers tr on tr.TRIGGERID=f.TRIGGERID right join hosts h on i.HOSTID=h.HOSTID JOIN hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_='icmpping[,4]' and h.status=0 and t.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" and tr.STATUS=0 group by i.ITEMID,h.name,f.TRIGGERID order by h.HOSTID";
	# cursor.execute(query)
	# icmp = cursor.fetchall()
	
	query="select h.HOSTID,round(AVG(his.VALUE_AVG),3) from trends his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.name in ('CPU Utilization') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	cpu_sys = cursor.fetchall()
	query="select h.HOSTID,round(max(his.value_max),3) from trends his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.name in ('CPU Utilization') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	cpu_sys_max = cursor.fetchall()
	# query="select h.HOSTID,round(AVG(his.VALUE_AVG),3) from trends his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('system.cpu.util[,system]','system.stat[cpu,sy]','system.cpu.util[,,avg5]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	# cursor.execute(query)
	# cpu_sys = cursor.fetchall()

	# query="select h.HOSTID,round(AVG(his.VALUE_AVG),3) from trends his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('system.cpu.util[,nice]','system.stat[cpu,ni]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	# cursor.execute(query)
	# cpu_ni = cursor.fetchall()
	# query="select h.HOSTID,round(AVG(his.VALUE_AVG),3) from trends his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('system.cpu.util[,user]','system.stat[cpu,us]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	# cursor.execute(query)
	# cpu_us = cursor.fetchall()

	query="select h.HOSTID,round(AVG(his.VALUE_AVG),3)  from trends his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('vm.memory.size[pused]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	memory_pused = cursor.fetchall()
	query="select h.HOSTID,round(AVG(his.VALUE_AVG),3)  from trends_uint his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('vm.memory.size[free]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	memory_free = cursor.fetchall()
	query="select h.HOSTID,round(AVG(his.VALUE_AVG),3)  from trends_uint his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('vm.memory.size[available]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	memory_available = cursor.fetchall()
	
	query="select h.HOSTID,round(AVG(his.VALUE_AVG),3) from trends_uint his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('vm.memory.size[total]') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	memory_total = cursor.fetchall()
	query="select h.NAME,h.HOSTID,i.KEY_,round(AVG(his.VALUE_AVG),3) as avg from trends_uint his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ like ('%vfs.fs.size%') and i.KEY_ like ('%used%') and i.KEY_ not like '%FSNAME%' and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,i.KEY_ order by h.HOSTID,i.KEY_";
	query="select h.NAME,h.HOSTID,i.KEY_,his.VALUE_AVG,max(his.clock),his.clock from trends_uint his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ like ('%vfs.fs.size%') and i.KEY_ like ('%used%') and i.KEY_ not like '%FSNAME%' and h.status=0 and his.clock BETWEEN "+str(round(time.time()-86400))+" and "+str(round(time.time()))+" group by h.NAME,i.KEY_ order by h.HOSTID,i.KEY_";
	cursor.execute(query)
	disks_used = cursor.fetchall()
	disk_used_dict=dict()
	for disk_used in disks_used:
		search_found=False
		for s in consider_disks:
			search_found=(disk_used[2][12:-6].__contains__(s) or disk_used[2][12:-6].__contains__(":")) and (not(disk_used[2][12:-6].__contains__("C:")))
			if search_found:
				break
		if(disk_used_dict.get(disk_used[1],0) == 0):
			data={disk_used[1]:0}
			disk_used_dict.update(data)
		if search_found:
			disk_used_dict[disk_used[1]] +=disk_used[3]
	query="select h.NAME,h.HOSTID,i.KEY_,his.VALUE_AVG,max(his.clock) from trends_uint his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ like ('%vfs.fs.size%') and i.KEY_ like ('%total%') and i.KEY_ not like '%FSNAME%' and h.status=0 and his.clock BETWEEN "+str(round(time.time()-86400))+" and "+str(round(time.time()))+" group by h.NAME,i.KEY_,his.VALUE_AVG order by h.HOSTID,i.KEY_";
	
	cursor.execute(query)
	disks_total = cursor.fetchall()
	disk_total_dict=dict()
	for disk_total in disks_total:
		search_found=False
		for s in consider_disks:
			#search_found=disk_total[2][12:-7].__contains__(s)
			search_found=(disk_total[2][12:-7].__contains__(s) or disk_total[2][12:-7].__contains__(":")) and (not(disk_total[2][12:-7].__contains__("C:")))
			if search_found:
				break
		if(disk_total_dict.get(disk_total[1],0) == 0):
			data={disk_total[1]:0}
			disk_total_dict.update(data)
		if search_found:
			disk_total_dict[disk_total[1]] +=disk_total[3]
	query="select h.HOSTID,value from history_str his  join items i on his.itemid=i.itemid join hosts h on i.hostid=h.hostid join hosts_groups hg on hg.HOSTID=h.HOSTID where hg.GROUPID="+str(groupids)+" and i.KEY_ in ('system.hostname') and h.status=0 and his.clock BETWEEN "+str(from_time)+" and "+str(to_time)+" group by h.NAME,h.HOSTID order by h.HOSTID";
	cursor.execute(query)
	host_name = cursor.fetchall()
	cpu_sys_dict=dict(cpu_sys)
	cpu_sys_max_dict=dict(cpu_sys_max)
	#cpu_ni_dict=dict(cpu_ni)
	#cpu_us_dict=dict(cpu_us)
	memory_free_dict=dict(memory_free)
	memory_available_dict=dict(memory_available)
	memory_pused_dict=dict(memory_pused)
	memory_total_dict=dict(memory_total)
	host_name_dict=dict(host_name)
	ips_dict=dict(ips)

	for host in hosts:
		data = dict()
		data['Zabbix Host Name'] = host[1]
		data['IP']=str(ips_dict.get(host[0],"0.0.0.0"))
		data['Host Name']=str(host_name_dict.get(host[0],"N/A"))
		data['CPU Usage (Average %)']=str(round(cpu_sys_dict.get(host[0],0),2))+'%'
		data['CPU Usage (Max %)']=str(round(cpu_sys_max_dict.get(host[0],0),2))+'%'
		#data['RAM Utilization (%)']=str(round(((memory_total_dict.get(host[0],0)-memory_free_dict.get(host[0],0))*100)/(memory_total_dict.get(host[0],1)),2))+'%'
		data['Free Utilization (%)']=str(round(((memory_total_dict.get(host[0],0)-memory_free_dict.get(host[0],0))*100)/(memory_total_dict.get(host[0],1)),2))+'%'
		if(host[0] in memory_available_dict):
			data['Available Utilization (%)']=str(round(((memory_total_dict.get(host[0],0)-memory_available_dict.get(host[0],0))*100)/(memory_total_dict.get(host[0],1)),2))+'%'
		else:
			data['Available Utilization (%)']="N/A"
		data['RAM Utilization (%)']=str(round(memory_pused_dict.get(host[0],0),2))+'%'
		data['RAM Total (GB)']=round((((memory_total_dict.get(host[0],0)/1024)/1024)/1024),2)
		#data['Storage: Used (GB)']=str(disk_used_dict.get(host[0],0))+'GB'
		#data['Storage: Reserved (GB)']=str(disk_total_dict.get(host[0],0))+'GB'
		data['Storage: Used (GB)']=round((((disk_used_dict.get(host[0],0)/1024)/1024)/1024),2)
		data['Storage: Reserved (GB)']=round((((disk_total_dict.get(host[0],0)/1024)/1024)/1024),2)
		hostsinfo.append(data)

	wb = Workbook()
	ws = wb.active
	#ws.column_dimensions=100
	ws.column_dimensions['A'].width = 40
	ws.column_dimensions['B'].width = 20
	ws.column_dimensions['C'].width = 20
	ws.column_dimensions['D'].width = 20
	ws.column_dimensions['E'].width = 20
	ws.column_dimensions['F'].width = 20
	ws.column_dimensions['G'].width = 20
	ws.column_dimensions['H'].width = 20
	ws.merge_cells('A1:E1')
	top_left_cell = ws['A1']
	top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
	ws['A1'] = "Storage"
	ws.append(['','','Calendar month',str(datetime.fromtimestamp(from_time).month)+'.'+str(datetime.fromtimestamp(from_time).year),str(datetime.fromtimestamp(from_time).month)+'.'+str(datetime.fromtimestamp(from_time).year)])
	header=["Zabbix Host Name","IP","Host Name","Storage: Reserved (GB)","Storage: Used (GB)"]
	ws.append(header)
	for hostinfo in hostsinfo:
		values=[hostinfo.get("Zabbix Host Name"),hostinfo.get("IP"),hostinfo.get("Host Name"),hostinfo.get("Storage: Reserved (GB)"),hostinfo.get("Storage: Used (GB)")]
		ws.append(values)
	#ws.insert_rows(ws._current_row)
	header=["Zabbix Host Name","IP","Host Name","RAM Total (GB)","RAM Utilization (without cache) (%)","RAM Utilization (with cache) (%)","CPU Usage (Average %)","CPU Usage (Max %)"]
	#ws.insert_rows(ws._current_row+1)
	ws.append(["\n"])
	ws.append(["CPU & Memory utilization"])
	ws.merge_cells('A'+str(ws._current_row)+':'+'F'+str(ws._current_row))
	top_left_cell = ws['A'+str(ws._current_row)]
	top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
	ws.append(['','','Calendar month','',str(datetime.fromtimestamp(from_time).month)+'.'+str(datetime.fromtimestamp(from_time).year),str(datetime.fromtimestamp(from_time).month)+'.'+str(datetime.fromtimestamp(from_time).year)])
	ws.append(header)
	for hostinfo in hostsinfo:
		values=[hostinfo.get("Zabbix Host Name"),hostinfo.get("IP"),hostinfo.get("Host Name"),hostinfo.get("RAM Total (GB)"),hostinfo.get("RAM Utilization (%)"),hostinfo.get("Available Utilization (%)"),hostinfo.get("CPU Usage (Average %)"),hostinfo.get("CPU Usage (Max %)")]
		ws.append(values)
	file_name = groupname+".csv"
	wb.save(dir_name+'/'+file_name)
	print(file_name)
count=0
with open(file_name,"wt+") as f:
	 for hostinfo in hostsinfo:
		 writer = csv.writer(f,lineterminator='\n')	
		 if(count == 0 ):
			 header=["Zabbix Host Name","IP","Host Name","Storage: Reserved (GB)","Storage: Used (GB)"]
			 writer.writerow(header)
		 values=[hostinfo.get("Zabbix Host Name"),hostinfo.get("IP"),hostinfo.get("Host Name"),hostinfo.get("Storage: Reserved (GB)"),hostinfo.get("Storage: Used (GB)")]
		 writer.writerow(values)
		 count=count+1
		 print(hostinfo)
	 writer = csv.writer(f,lineterminator='\n')
	 writer.writerow("\n")
	 count=0
	 for hostinfo in hostsinfo:
		 writer = csv.writer(f,lineterminator='\n')	
		 if(count == 0 ):
			 header=["Zabbix Host Name","IP","Host Name","RAM Total (GB)","RAM Utilization (with cache) (%)","RAM Utilization (without cache) (%)","CPU Usage (Average %)","CPU Usage (Max %)"]
			 writer.writerow(header)
		 values=[hostinfo.get("Zabbix Host Name"),hostinfo.get("IP"),hostinfo.get("Host Name"),hostinfo.get("RAM Total (GB)"),hostinfo.get("RAM Utilization (%)"),hostinfo.get("Available Utilization (%)"),hostinfo.get("CPU Usage (Average %)"),hostinfo.get("CPU Usage (Max %)")]
		 writer.writerow(values)
		 count=count+1
		 print(hostinfo)
	 writer = csv.writer(f,lineterminator='\n')
	 writer.writerow("\n")
f.close()			
